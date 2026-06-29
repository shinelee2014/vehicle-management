"""
进出记录 API（核心）
"""
from datetime import datetime, timedelta
import os
from fastapi import APIRouter, Depends, HTTPException, Request, Query
from fastapi.responses import FileResponse
from sqlalchemy import or_, and_, func
from sqlalchemy.orm import Session
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.database import get_db
from app.core.permissions import get_current_user
from app.config import settings
from app.models.user import User, UserRole
from app.models.post import Post
from app.models.record import (
    Record, VehicleType, Direction, ApprovalStatus, RecordStatus
)
from app.models.message import Message, MessageType
from app.schemas.record import RecordCreateIn, RecordCreateOut, ApprovalRequest
from app.services.audit import log_audit

router = APIRouter()


def _gen_record_no(direction: str) -> str:
    """生成业务编号：IN-20260623-001"""
    prefix = "IN" if direction == "in" else "OUT"
    date_str = datetime.now().strftime("%Y%m%d")
    return f"{prefix}-{date_str}-{int(datetime.now().timestamp()) % 100000:05d}"


def _safe_val(v):
    """安全获取枚举或字符串的值"""
    if hasattr(v, 'value'):
        return v.value
    return v


def _send_approval_message(db: Session, recipient_id: int, record: Record) -> None:
    """发送待审批通知"""
    vehicle = _safe_val(record.vehicle_type)
    direction = _safe_val(record.direction)
    msg = Message(
        recipient_id=recipient_id,
        sender_id=record.operator_id,
        sender_name=record.operator_name,
        type=MessageType.APPROVAL_REQUEST,
        title=f"【待审批】{record.plate_number} {vehicle} {'进场' if direction == 'in' else '出场'}",
        content=(
            f"操作人：{record.operator_name}\n"
            f"车牌：{record.plate_number}\n"
            f"类型：{vehicle}\n"
            f"事由：{record.in_remark or record.out_remark or '无'}\n"
            f"货物：{record.cargo_info or '无'}\n"
            f"时间：{(record.in_time or record.out_time).strftime('%Y-%m-%d %H:%M:%S') if (record.in_time or record.out_time) else ''}"
        ),
        related_record_id=record.id,
    )
    db.add(msg)


@router.post("/in")
async def create_in(
    req: RecordCreateIn,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """提交进场登记"""
    # 检查车辆类型是否合法且启用
    from app.models.vehicle_type import VehicleType
    vt = db.query(VehicleType).filter(VehicleType.code == req.vehicle_type, VehicleType.is_active == True).first()
    if not vt:
        raise HTTPException(status_code=400, detail={"code": 400, "message": f"未知的车辆类型或该车型已被禁用: {req.vehicle_type}"})

    # 检查审批人
    approver = db.query(User).filter(User.id == req.approver_id, User.is_approver == True, User.is_active == True).first()
    if not approver:
        raise HTTPException(status_code=400, detail={"code": 400, "message": "审批人不存在或无审批权限"})

    # 检查岗亭
    post = db.query(Post).filter(Post.id == req.post_id, Post.is_active == True).first()
    if not post:
        raise HTTPException(status_code=400, detail={"code": 400, "message": "岗亭不存在"})

    now = datetime.utcnow()
    record = Record(
        record_no=_gen_record_no("in"),
        plate_number=req.plate_number.upper(),
        vehicle_type=req.vehicle_type,
        direction=Direction.IN.value,
        post_id=req.post_id,
        operator_id=current_user.id,
        operator_name=current_user.real_name,
        in_time=now,
        in_photos=[p.dict() for p in req.photos],
        in_remark=req.in_remark,
        cargo_info=req.cargo_info if req.vehicle_type == "truck" else None,
        approver_id=req.approver_id,
        approver_name=approver.real_name,
        approval_status=ApprovalStatus.PENDING.value,
        status=RecordStatus.IN_PENDING.value,
    )
    db.add(record)
    db.flush()

    # 通知审批人
    _send_approval_message(db, req.approver_id, record)

    log_audit(db, user_id=current_user.id, username=current_user.username, action="create_in_record",
              target_type="record", target_id=record.id, details={"plate": req.plate_number},
              ip_address=request.client.host if request.client else "")
    db.commit()

    return {"code": 0, "message": "提交成功，已通知审批人", "data": record.to_dict()}


@router.post("/out")
async def create_out(
    req: RecordCreateOut,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """提交出场登记"""
    # 检查审批人
    approver = db.query(User).filter(User.id == req.approver_id, User.is_approver == True, User.is_active == True).first()
    if not approver:
        raise HTTPException(status_code=400, detail={"code": 400, "message": "审批人不存在或无审批权限"})

    # 检查岗亭
    post = db.query(Post).filter(Post.id == req.post_id, Post.is_active == True).first()
    if not post:
        raise HTTPException(status_code=400, detail={"code": 400, "message": "岗亭不存在"})

    now = datetime.utcnow()

    # 关联进记录
    related = None
    if req.related_record_id:
        related = db.query(Record).filter(Record.id == req.related_record_id).first()
    else:
        # 找最近一条同车牌的、已通过进场的记录
        related = (
            db.query(Record)
            .filter(
                Record.plate_number == req.plate_number.upper(),
                Record.direction == Direction.IN.value,
                Record.approval_status.in_([ApprovalStatus.APPROVED.value, ApprovalStatus.TIMEOUT.value]),
                Record.companion_id.is_(None),
            )
            .order_by(Record.in_time.desc())
            .first()
        )
    if not related:
        raise HTTPException(status_code=400, detail={"code": 400, "message": "找不到对应的进记录，请先扫码或搜车牌"})

    # 创建出记录
    out_record = Record(
        record_no=_gen_record_no("out"),
        plate_number=req.plate_number.upper(),
        vehicle_type=related.vehicle_type if isinstance(related.vehicle_type, str) else related.vehicle_type.value,
        direction=Direction.OUT.value,
        post_id=req.post_id,
        operator_id=current_user.id,
        operator_name=current_user.real_name,
        out_time=now,
        out_photos=[p.dict() for p in req.photos],
        out_remark=req.out_remark,
        loading_start_at=related.in_time,
        loading_end_at=now,
        loading_duration=int((now - related.in_time).total_seconds() / 60) if related.in_time else None,
        approver_id=req.approver_id,
        approver_name=approver.real_name,
        approval_status=ApprovalStatus.PENDING.value,
        status=RecordStatus.OUT_PENDING.value,
        related_record_id=related.id,
    )
    db.add(out_record)
    db.flush()

    # 关联双方
    related.companion_id = out_record.id
    out_record.companion_id = related.id

    # 通知审批人
    _send_approval_message(db, req.approver_id, out_record)

    log_audit(db, user_id=current_user.id, username=current_user.username, action="create_out_record",
              target_type="record", target_id=out_record.id, details={"plate": req.plate_number, "related": related.id},
              ip_address=request.client.host if request.client else "")
    db.commit()

    return {"code": 0, "message": "提交成功，已通知审批人", "data": out_record.to_dict()}


@router.get("/")
async def list_records(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    plate_number: Optional[str] = None,
    post_id: Optional[int] = None,
    vehicle_type: Optional[str] = None,
    approval_status: Optional[str] = None,
    direction: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    keyword: Optional[str] = None,
    include_deleted: bool = Query(False, description="admin 用：包含已软删除的记录"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """记录列表（按条件筛选）"""
    query = db.query(Record)

    # 默认隐藏已删除，只有 admin 显式 include_deleted=True 才显示
    if not include_deleted:
        query = query.filter(Record.is_deleted == 0)
    else:
        # 非 admin 不能看已删记录
        if current_user.role != UserRole.ADMIN.value:
            raise HTTPException(status_code=403, detail={"code": 403, "message": "仅管理员可查看已删除记录"})

    if plate_number:
        query = query.filter(Record.plate_number.like(f"%{plate_number.upper()}%"))
    if post_id:
        query = query.filter(Record.post_id == post_id)
    if vehicle_type:
        query = query.filter(Record.vehicle_type == vehicle_type)
    if approval_status:
        query = query.filter(Record.approval_status == approval_status)
    if direction:
        query = query.filter(Record.direction == direction)
    if start_date:
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d")
            query = query.filter(or_(Record.in_time >= start, Record.out_time >= start))
        except ValueError:
            pass
    if end_date:
        try:
            end = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
            query = query.filter(or_(Record.in_time < end, Record.out_time < end))
        except ValueError:
            pass

    # 按角色过滤
    if current_user.role == UserRole.SECURITY.value:
        # 保安只能看自己的
        query = query.filter(Record.operator_id == current_user.id)

    query = query.order_by(Record.created_at.desc())
    total = query.count()
    items = query.offset((page - 1) * page_size).limit(page_size).all()

    return {
        "code": 0,
        "message": "ok",
        "data": {
            "items": [r.to_dict() for r in items],
            "total": total,
            "page": page,
            "page_size": page_size,
        },
    }


@router.get("/pending")
async def list_pending(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """待我审批列表"""
    items = (
        db.query(Record)
        .filter(
            Record.approver_id == current_user.id,
            Record.approval_status == ApprovalStatus.PENDING,
        )
        .order_by(Record.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    total = (
        db.query(func.count(Record.id))
        .filter(
            Record.approver_id == current_user.id,
            Record.approval_status == ApprovalStatus.PENDING,
        )
        .scalar()
    )
    return {
        "code": 0,
        "message": "ok",
        "data": {
            "items": [r.to_dict() for r in items],
            "total": total,
            "page": page,
            "page_size": page_size,
        },
    }


@router.get("/unbilled-list")
async def unbilled_list(
    plate_number: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """未完成出场列表（出场时搜车牌关联用）"""
    query = (
        db.query(Record)
        .filter(
            Record.direction == Direction.IN,
            Record.approval_status.in_([ApprovalStatus.APPROVED, ApprovalStatus.TIMEOUT]),
            Record.companion_id.is_(None),
        )
    )
    if plate_number:
        query = query.filter(Record.plate_number.like(f"%{plate_number.upper()}%"))
    items = query.order_by(Record.in_time.desc()).limit(50).all()
    return {
        "code": 0,
        "message": "ok",
        "data": [r.to_dict() for r in items],
    }


@router.get("/{record_id}")
async def get_record(
    record_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """记录详情"""
    record = db.query(Record).filter(Record.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail={"code": 404, "message": "记录不存在"})
    return {"code": 0, "message": "ok", "data": record.to_dict()}


@router.post("/{record_id}/approve")
async def approve(
    record_id: int,
    req: ApprovalRequest,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """审批通过"""
    if req.action != "approve":
        raise HTTPException(status_code=400, detail={"code": 400, "message": "操作类型错误"})

    record = db.query(Record).filter(Record.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail={"code": 404, "message": "记录不存在"})

    if record.approver_id != current_user.id:
        raise HTTPException(status_code=403, detail={"code": 403, "message": "您不是该记录的审批人"})

    if record.approval_status != ApprovalStatus.PENDING.value:
        raise HTTPException(status_code=400, detail={"code": 400, "message": "该记录已审批"})

    now = datetime.utcnow()
    record.approval_status = ApprovalStatus.APPROVED.value
    record.approval_time = now
    record.approval_remark = req.remark

    if record.direction == Direction.IN.value:
        record.status = RecordStatus.IN_APPROVED.value
    else:
        record.status = RecordStatus.OUT_APPROVED.value
        # 出场通过后，如果有关联进记录，也标记完成
        if record.related_record_id:
            related = db.query(Record).filter(Record.id == record.related_record_id).first()
            if related:
                # 进记录状态保持 IN_APPROVED，不改成 completed（避免破坏流程）
                pass

    # 通知保安（审批结果）
    msg = Message(
        recipient_id=record.operator_id,
        sender_id=current_user.id,
        sender_name=current_user.real_name,
        type=MessageType.APPROVAL_RESULT,
        title=f"【已通过】{record.plate_number} {'进场' if record.direction == Direction.IN.value else '出场'}",
        content=f"审批人：{current_user.real_name}\n审批意见：{req.remark or '无'}",
        related_record_id=record.id,
    )
    db.add(msg)

    log_audit(db, user_id=current_user.id, username=current_user.username, action="approve",
              target_type="record", target_id=record.id, ip_address=request.client.host if request.client else "")
    db.commit()
    return {"code": 0, "message": "已通过", "data": record.to_dict()}


@router.post("/{record_id}/reject")
async def reject(
    record_id: int,
    req: ApprovalRequest,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """审批驳回"""
    if req.action != "reject":
        raise HTTPException(status_code=400, detail={"code": 400, "message": "操作类型错误"})

    record = db.query(Record).filter(Record.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail={"code": 404, "message": "记录不存在"})

    if record.approver_id != current_user.id:
        raise HTTPException(status_code=403, detail={"code": 403, "message": "您不是该记录的审批人"})

    if record.approval_status != ApprovalStatus.PENDING.value:
        raise HTTPException(status_code=400, detail={"code": 400, "message": "该记录已审批"})

    now = datetime.utcnow()
    record.approval_status = ApprovalStatus.REJECTED.value
    record.approval_time = now
    record.approval_remark = req.remark or "驳回"

    if record.direction == Direction.IN.value:
        record.status = RecordStatus.IN_REJECTED.value
    else:
        record.status = RecordStatus.OUT_REJECTED.value

    # 通知保安
    msg = Message(
        recipient_id=record.operator_id,
        sender_id=current_user.id,
        sender_name=current_user.real_name,
        type=MessageType.APPROVAL_RESULT,
        title=f"【已驳回】{record.plate_number} {'进场' if record.direction == Direction.IN.value else '出场'}",
        content=f"审批人：{current_user.real_name}\n驳回原因：{req.remark or '无'}",
        related_record_id=record.id,
    )
    db.add(msg)

    log_audit(db, user_id=current_user.id, username=current_user.username, action="reject",
              target_type="record", target_id=record.id, ip_address=request.client.host if request.client else "")
    db.commit()
    return {"code": 0, "message": "已驳回", "data": record.to_dict()}


@router.get("/export/excel")
async def export_excel(
    plate_number: Optional[str] = None,
    post_id: Optional[int] = None,
    vehicle_type: Optional[str] = None,
    approval_status: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    request: Request = None,
):
    """导出 Excel"""
    if current_user.role == UserRole.SECURITY.value:
        raise HTTPException(status_code=403, detail={"code": 403, "message": "无导出权限"})

    query = db.query(Record)
    if plate_number:
        query = query.filter(Record.plate_number.like(f"%{plate_number.upper()}%"))
    if post_id:
        query = query.filter(Record.post_id == post_id)
    if vehicle_type:
        query = query.filter(Record.vehicle_type == vehicle_type)
    if approval_status:
        query = query.filter(Record.approval_status == approval_status)
    if start_date:
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d")
            query = query.filter(or_(Record.in_time >= start, Record.out_time >= start))
        except ValueError:
            pass
    if end_date:
        try:
            end = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
            query = query.filter(or_(Record.in_time < end, Record.out_time < end))
        except ValueError:
            pass

    items = query.order_by(Record.created_at.desc()).limit(5000).all()

    # 生成 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "车辆进出记录"

    # 表头
    headers = ["业务编号", "车牌", "类型", "方向", "岗亭", "操作人", "进场时间", "出场时间", "停留(分)", "货物", "事由", "审批人", "审批状态", "审批时间", "审批意见"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="409EFF", end_color="409EFF", fill_type="solid")

    # 岗亭映射
    posts = {p.id: p.name for p in db.query(Post).all()}

    # 数据
    for r in items:
        vtype = r.vehicle_type.value if hasattr(r.vehicle_type, "value") else r.vehicle_type
        direction = "进场" if (r.direction.value if hasattr(r.direction, "value") else r.direction) == "in" else "出场"
        ws.append([
            r.record_no,
            r.plate_number,
            {"internal": "内部车", "external": "外部车", "truck": "货车"}.get(vtype, vtype),
            direction,
            posts.get(r.post_id, ""),
            r.operator_name,
            r.in_time.strftime("%Y-%m-%d %H:%M:%S") if r.in_time else "",
            r.out_time.strftime("%Y-%m-%d %H:%M:%S") if r.out_time else "",
            r.loading_duration or "",
            r.cargo_info or "",
            r.in_remark or r.out_remark or "",
            r.approver_name or "",
            {"pending": "待审批", "approved": "已通过", "rejected": "已驳回", "timeout": "已超时"}.get(
                r.approval_status.value if hasattr(r.approval_status, "value") else r.approval_status, ""
            ),
            r.approval_time.strftime("%Y-%m-%d %H:%M:%S") if r.approval_time else "",
            r.approval_remark or "",
        ])

    # 列宽
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)

    # 保存
    filename = f"车辆记录_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join("/tmp", filename)
    os.makedirs("/tmp", exist_ok=True)
    wb.save(filepath)

    log_audit(db, user_id=current_user.id, username=current_user.username, action="export",
              target_type="record", ip_address=request.client.host if (request and request.client) else "")
    db.commit()

    return FileResponse(
        filepath,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )


@router.post("/batch-delete")
async def batch_delete_records(
    payload: dict,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """批量软删除记录（仅 admin）

    body: {record_ids: [1,2,3], reason: "测试数据清理"}

    行为：
    - 已删过的、状态不允许删的会跳过（返回 skipped 列表）
    - 默认拒绝删除已审批通过且未撤销的记录（force=true 才删）
    - 删除前把每条记录的完整快照写到 audit_logs.details（防止真删后无法恢复）
    - 列表默认隐藏已删记录
    """
    if current_user.role != UserRole.ADMIN.value:
        raise HTTPException(status_code=403, detail={"code": 403, "message": "仅管理员可删除记录"})

    record_ids = payload.get("record_ids") or []
    reason = (payload.get("reason") or "").strip()
    force = bool(payload.get("force", False))

    if not isinstance(record_ids, list) or not record_ids:
        raise HTTPException(status_code=400, detail={"code": 400, "message": "record_ids 必须是非空数组"})
    if not reason:
        raise HTTPException(status_code=400, detail={"code": 400, "message": "删除原因必填（必填，便于审计）"})
    if len(record_ids) > 500:
        raise HTTPException(status_code=400, detail={"code": 400, "message": "单次最多删除 500 条"})
    if len(reason) > 200:
        raise HTTPException(status_code=400, detail={"code": 400, "message": "原因不超过 200 字"})

    now = datetime.utcnow()
    deleted = []
    skipped = []

    for rid in record_ids:
        try:
            rid_int = int(rid)
        except (TypeError, ValueError):
            skipped.append({"id": rid, "reason": "非数字 id"})
            continue

        r = db.query(Record).filter(Record.id == rid_int).first()
        if not r:
            skipped.append({"id": rid_int, "reason": "记录不存在"})
            continue
        if r.is_deleted:
            skipped.append({"id": rid_int, "reason": "已删除"})
            continue

        # 保护：已审批通过 + 已出场 + 是进记录的核心数据，不让随便删
        approval = r.approval_status.value if hasattr(r.approval_status, "value") else r.approval_status
        direction = r.direction.value if hasattr(r.direction, "value") else r.direction
        is_locked = (approval == "approved" and direction == "in" and r.companion_id is not None)
        if is_locked and not force:
            skipped.append({"id": rid_int, "reason": "已审批 + 已匹配出场，需 force=true 才删"})
            continue

        # 写审计快照（删之前）
        snapshot = {
            "id": r.id,
            "record_no": r.record_no,
            "plate_number": r.plate_number,
            "vehicle_type": r.vehicle_type.value if hasattr(r.vehicle_type, "value") else r.vehicle_type,
            "direction": r.direction.value if hasattr(r.direction, "value") else r.direction,
            "post_id": r.post_id,
            "operator_id": r.operator_id,
            "operator_name": r.operator_name,
            "in_time": r.in_time.isoformat() if r.in_time else None,
            "in_remark": r.in_remark,
            "out_time": r.out_time.isoformat() if r.out_time else None,
            "out_remark": r.out_remark,
            "cargo_info": r.cargo_info,
            "approver_name": r.approver_name,
            "approval_status": approval,
            "approval_time": r.approval_time.isoformat() if r.approval_time else None,
            "approval_remark": r.approval_remark,
            "related_record_id": r.related_record_id,
            "companion_id": r.companion_id,
            "status": r.status.value if hasattr(r.status, "value") else r.status,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        }

        r.is_deleted = 1
        r.deleted_at = now
        r.deleted_by = current_user.id
        r.deleted_by_name = current_user.real_name
        r.deleted_reason = reason

        log_audit(
            db, user_id=current_user.id, username=current_user.username,
            action="soft_delete_record", target_type="record", target_id=r.id,
            details={"reason": reason, "force": force, "snapshot": snapshot},
            ip_address=request.client.host if request.client else "",
        )
        deleted.append({"id": r.id, "record_no": r.record_no})

    db.commit()

    return {
        "code": 0,
        "message": f"删除 {len(deleted)} 条，跳过 {len(skipped)} 条",
        "data": {"deleted": deleted, "skipped": skipped, "deleted_count": len(deleted), "skipped_count": len(skipped)},
    }


@router.post("/{record_id}/restore")
async def restore_record(
    record_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """恢复已软删除的记录（仅 admin）"""
    if current_user.role != UserRole.ADMIN.value:
        raise HTTPException(status_code=403, detail={"code": 403, "message": "仅管理员可恢复记录"})

    r = db.query(Record).filter(Record.id == record_id).first()
    if not r:
        raise HTTPException(status_code=404, detail={"code": 404, "message": "记录不存在"})
    if not r.is_deleted:
        raise HTTPException(status_code=400, detail={"code": 400, "message": "记录未被删除，无需恢复"})

    r.is_deleted = 0
    r.deleted_at = None
    r.deleted_by = None
    r.deleted_by_name = None
    r.deleted_reason = None

    log_audit(
        db, user_id=current_user.id, username=current_user.username,
        action="restore_record", target_type="record", target_id=r.id,
        ip_address=request.client.host if request.client else "",
    )
    db.commit()
    return {"code": 0, "message": "已恢复", "data": r.to_dict()}
